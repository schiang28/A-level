import openpyxl as px
from openpyxl.styles import PatternFill

wbname = "Example.xlsx"
wb = px.load_workbook(wbname, read_only=True)
ws = wb.active

new_wb = px.Workbook()
new_ws = new_wb.create_sheet(title="I Made This")
new_ws.sheet_view.showGridLines = False
new_ws.sheet_state = "visible"
del new_wb["Sheet"]

for r, row in enumerate(ws.iter_rows(min_row=2), start=1):
    name, year, reg, adno, house = row

    second, first = name.value.split(", ")
    tutor = reg.value.split("-")[1]
    house = house.value.split("-")[1]

    for c, value in enumerate([adno.value, first, second, tutor, house], start=1):
        cell = new_ws.cell(column=c, row=r)
        cell.value = value
        cell.fill = PatternFill(
            start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid"
        )

new_wb.save("Written.xlsx")
